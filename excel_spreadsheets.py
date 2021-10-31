import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import os
import sys
import programs_code


def unwantsignsremove(strng=''):
    removethose = '().,-_*'
    for x in removethose:
        strng = strng.replace(x, ' ')
    return strng


def xl_spreadsheets(filename):
    wb = xl.load_workbook(filename, read_only=False)
    sheet = wb['Лист1']
    # cell = sheet['a1']
    # print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        discount_price = cell.value * 0.9
        discount_price_cell = sheet.cell(row, 4)
        discount_price_cell.value = discount_price

    values = Reference(sheet, max_row=sheet.max_row, min_col=4, min_row=2, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'a6')

    wb.save(filename)


# Extracting the signal list out of the Eplan signals export file to another Excel workbook
def epl2signal_list(filename):
    if not os.path.isfile(filename):
        print(f"The File {filename} is not found!")
        return
    if '.xlsx' not in filename:
        print('File needs to be in .xlsx format!')
        return
    print('Processing input data...')
    wbook = xl.load_workbook(filename, read_only=False)
    epl_sheet = wbook['EplSheet']
#   check for existence workbook 'generated_list'
    if 'generated_list' in wbook.sheetnames:
        print('Rewriting generated_list...')
        del wbook['generated_list']
    else:
        print('Creating new sheet...')
    list_sheet = wbook.create_sheet(title='generated_list')
    new_list_row = 1
    old_epl_plcmodule_cell_value = ''
    for row in range(3, epl_sheet.max_row + 1):
        epl_signum_cell = epl_sheet.cell(row, 6)
        epl_signame_cell = epl_sheet.cell(row, 7)
        epl_plcmodule_cell = epl_sheet.cell(row, 5)
        new_signum_cell = list_sheet.cell(new_list_row, 2)
        new_signame_cell = list_sheet.cell(new_list_row, 4)
        new_plcmodule_cell = list_sheet.cell(new_list_row, 3)
        open_commentbracket_cell = list_sheet.cell(new_list_row, 1)
        close_commentbracket_cell = list_sheet.cell(new_list_row, 9)
        #defining the type of signals in list
        if old_epl_plcmodule_cell_value != epl_plcmodule_cell.value:
            if 'DI' in epl_plcmodule_cell.value and 'DI' not in old_epl_plcmodule_cell_value:
                open_commentbracket_cell.value = f'(*'
                close_commentbracket_cell.value = f'*)'
                new_signum_cell.value = '---------==== Входные дискретные сигналы ====---------'
                old_epl_plcmodule_cell_value = epl_plcmodule_cell.value
                new_list_row += 1
            elif 'DO' in epl_plcmodule_cell.value and 'DO' not in old_epl_plcmodule_cell_value:
                open_commentbracket_cell.value = f'(*'
                close_commentbracket_cell.value = f'*)'
                new_signum_cell.value = '---------==== Выходные дискретные сигналы ====---------'
                old_epl_plcmodule_cell_value = epl_plcmodule_cell.value
                new_list_row += 1
            elif 'AI' in epl_plcmodule_cell.value and 'AI' not in old_epl_plcmodule_cell_value:
                open_commentbracket_cell.value = f'(*'
                close_commentbracket_cell.value = f'*)'
                new_signum_cell.value = '---------==== Входные аналоговые параметры ====---------'
                old_epl_plcmodule_cell_value = epl_plcmodule_cell.value
                new_list_row += 1
            elif 'AO' in epl_plcmodule_cell.value and 'AO' not in old_epl_plcmodule_cell_value:
                open_commentbracket_cell.value = f'(*'
                close_commentbracket_cell.value = f'*)'
                new_signum_cell.value = '---------==== Выходные аналоговые параметры ====---------'
                old_epl_plcmodule_cell_value = epl_plcmodule_cell.value
                new_list_row += 1
        if epl_signum_cell.value is not None:
            new_list_row += 1
            #copy row with signal number field if it's not blank
            new_signum_cell.value = epl_signum_cell.value
            if epl_signame_cell.value is not None:
                new_signame_cell.value = unwantsignsremove(epl_signame_cell.value)
            else:
                new_signame_cell.value = epl_signame_cell.value
            new_plcmodule_cell.value = f'[{epl_plcmodule_cell.value}]'
            #printing comment brackets for the signal list strings
            open_commentbracket_cell.value = f'(*'
            close_commentbracket_cell.value = f'*)'
    print(f'Number of entries: {new_list_row - 1}')

    wbook.save(filename)
    wbook.close()
    print('File done!\n')


# Extracting the signal list out of the Eplan signals export file to .stf file
def epl2stf_signal_list(filename):
    if not os.path.isfile(filename):
        print(f"The File {filename} is not found!")
        return
    if '.xlsx' not in filename:
        print('File needs to be in .xlsx format!')
        return
    print('Processing input data...')
    wbook = xl.load_workbook(filename, read_only=False)
    epl_sheet = wbook['EplSheet']
#   check for existence
    if not os.path.isfile(filename):
        print('Creating .stf files...')
    else:
        print('Creating .stf file records...')
    old_epl_plcmodule_cell_value = ''
    signal_type = 0
    di_etries_counter = 0
    do_etries_counter = 0
    ai_etries_counter = 0
    di_signal_name_list = []
    ai_signal_name_list = []
    for row in range(3, epl_sheet.max_row + 1):
        epl_signum_cell = epl_sheet.cell(row, 6)
        epl_signame_cell = epl_sheet.cell(row, 7)
        epl_plcmodule_cell = epl_sheet.cell(row, 5)
        # defining the type of signals in list
        if old_epl_plcmodule_cell_value != epl_plcmodule_cell.value:
            if 'DI' in epl_plcmodule_cell.value and 'DI' not in old_epl_plcmodule_cell_value:
                di_file = open('DI.stf', "w")
                di_file.write(f'(* ---------==== Входные дискретные сигналы ====--------- *)\n')
                di_file.close()
                signal_type = 1
                old_epl_plcmodule_cell_value = epl_plcmodule_cell.value
            elif 'DO' in epl_plcmodule_cell.value and 'DO' not in old_epl_plcmodule_cell_value:
                do_file = open('DO.stf', "w")
                do_file.write(f'(* ---------==== Выходные дискретные сигналы ====--------- *)\n')
                do_file.close()
                signal_type = 2
                old_epl_plcmodule_cell_value = epl_plcmodule_cell.value
            elif 'AI' in epl_plcmodule_cell.value and 'AI' not in old_epl_plcmodule_cell_value:
                ai_file = open('AI.stf', "w")
                ai_file.write(f'(* ---------==== Входные аналоговые параметры ====--------- *)\n')
                ai_file.close()
                signal_type = 3
                old_epl_plcmodule_cell_value = epl_plcmodule_cell.value
#            elif 'AO' in epl_plcmodule_cell.value and 'AO' not in old_epl_plcmodule_cell_value:
#                new_signum_cell.value = '---------==== Выходные аналоговые параметры ====---------'
#                old_epl_plcmodule_cell_value = epl_plcmodule_cell.value
#                new_list_row += 1
            # else:
            #    print(f'Signal type choice error at entry: {row}')
        if epl_signum_cell.value is not None:
            # copy row with signal number field if it's not blank
            if signal_type == 1:
                di_file = open('DI.stf', "a")
                di_file.write(f'(*\t{epl_signum_cell.value}\t[{epl_plcmodule_cell.value}]\t{epl_signame_cell.value}\t*)\n')
                di_file.close()
                di_signal_name_list.append(epl_signame_cell.value)
                di_etries_counter += 1
            if signal_type == 2:
                do_file = open('DO.stf', "a")
                do_file.write(f'(*\t{epl_signum_cell.value}\t[{epl_plcmodule_cell.value}]\t{epl_signame_cell.value}\t*)\n')
                do_file.close()
                do_etries_counter += 1
            if signal_type == 3:
                ai_file = open('AI.stf', "a")
                ai_file.write(f'(*\t{epl_signum_cell.value}\t[{epl_plcmodule_cell.value}]\t{epl_signame_cell.value}\t*)\n')
                ai_file.close()
                ai_signal_name_list.append(epl_signame_cell.value)
                ai_etries_counter += 1

    # funcs call to write the rest of the Prog cade
    # print(di_signal_name_list)
    programs_code.input_discrete_code('DI.stf', di_etries_counter, di_signal_name_list)
    # programs_code.program_code('DO.stf', 2, do_etries_counter)
    programs_code.input_analogs_code('AI.stf', 3, ai_etries_counter)
    programs_code.input_analogs_init_code(ai_etries_counter, ai_signal_name_list)
    print('Files done!\n')


# xl_spreadsheets('transactions.xlsx')
epl2signal_list(sys.argv[1])      # need to find out how to transfer variable amount of arguments


#epl2stf_signal_list(sys.argv[1])

# print(unwantsignsremove('f,s(f.e)f*e-f_k'))







