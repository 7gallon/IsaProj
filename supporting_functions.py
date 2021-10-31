import os
import pyodbc
import openpyxl as xl


def make_ana_def(conf_filename, signals_filename):
    """
    Call this function right in the place where new DEF should be.
    """
    if not os.path.isfile(conf_filename):
        print(f"The File {conf_filename} is not found!")
        return
    if '.xlsx' not in conf_filename:
        print('File needs to be in .xlsx format!')
        return
    if not os.path.isfile(signals_filename):
        print(f"The File {signals_filename} is not found!")
        return
    if '.xlsx' not in signals_filename:
        print('File needs to be in .xlsx format!')
        return
    print('Processing generated analogs list for _DEF_ making...')
    conf_wb = xl.load_workbook(conf_filename, read_only=True)
    wb = xl.load_workbook(signals_filename, read_only=False)
    if 'ana_def_config' in conf_wb.sheetnames:
        config_sheet = conf_wb['ana_def_config']
        signals_sheet = wb['generated_list']
        print('Reading configuration...')
    else:
        print('No configuration found!')
        return
    analogs_signal_type_def_dictionary = {}
    analogs_signal_defs_values_dictionary = {}
    signtype_plcnum_dictionary = {'A1-AA': 'ANA1_', 'A2-AA': 'ANA2_', 'A13-AA': 'ANA3_', 'A4-AA': 'ANA4_',
                                  'A5-AA': 'ANA5_',
                                  'A1-ADI': 'DIS1_', 'A2-ADI': 'DIS2_', 'A13-ADI': 'DIS3_', 'A4-ADI': 'DIS4_',
                                  'A5-ADI': 'DIS5_',
                                  'A1-ADO': 'DO1_', 'A2-ADO': 'DO2_', 'A13-ADO': 'DO3_', 'A4-ADO': 'DO4_',
                                  'A5-ADO': 'DO5_'
                                  }
    # Filling up dictionaries out of ana_config workbook
    for row in range(1, config_sheet.max_row + 1):
        signal_type_cell = config_sheet.cell(row, 1)        # e.g. Температура
        signal_type_def_cell = config_sheet.cell(row, 2)    # e.g. T_
        analogs_signal_type_def_dictionary[signal_type_cell.value] = signal_type_def_cell.value
    # Comparing parsed signal names to dictionary keys
    for idx in range(2, signals_sheet.max_row + 1):
        signals_sheet_cell = signals_sheet.cell(idx, 4)
        plcnum_sheet_cell = signals_sheet.cell(idx, 3)
        varnum_sheet_cell = signals_sheet.cell(idx, 2)
        signal_name_frase = str(signals_sheet_cell.value)
        signal_name_words = signal_name_frase.split(" ")
        if plcnum_sheet_cell.value is not None:
            for typ in signtype_plcnum_dictionary:
                if typ in plcnum_sheet_cell.value:
                    analog_signal_plcnum = str(signtype_plcnum_dictionary[typ])
                #    print(typ, plcnum_sheet_cell.value)
                # else:
                #     print('NEED TO ADD MORE PLCS IN CODE!!!!!')
        else:
            # analog_signal_plcnum = '-------'
            continue
        result = 'DEF_' + analog_signal_plcnum
        for pos in signal_name_words:
            if pos in analogs_signal_type_def_dictionary:
                # print(pos)
                if analogs_signal_type_def_dictionary.get(pos) is not None and ('Резерв' not in pos):
                    result += str(analogs_signal_type_def_dictionary.get(pos)) + '_'
                else:
                    result = ''
        if result != '':
            analogs_signal_defs_values_dictionary[result[:-1:].upper()] = [varnum_sheet_cell.value, signal_name_frase]
            # print(result[:-1:].upper())
    return(analogs_signal_defs_values_dictionary)



def mdbedit(defsdictionary = {'1': ('2' ,'3')}):
    ''' Для работы pyodbc необходимо установить "AccessDatabaseEngine_X64,"`
        а для этого необходимо поставить х64 офис или хотябы удалить х32 '''
    # set up some constants
    MDB = 'd:/PrjLibrary.mdb'
    DRV = '{Microsoft Access Driver (*.mdb, *.accdb)}'
    # connect to db
    cnxn = pyodbc.connect('DRIVER={};DBQ={}'.format(DRV, MDB))
    crsr = cnxn.cursor()
    sql = "SELECT Name FROM Defines Where Define = '2' ; "  # your query goes here
    crsr.execute(sql)
    usednames = crsr.fetchall()
    print(usednames)
    for i in defsdictionary:
        tmpValuesList = defsdictionary.get(i)
        # tmpKeysList = defsdictionary.keys()
        print(i, tmpValuesList)
        if i not in usednames:
            crsr.execute("INSERT INTO Defines (Name, Define, SingleTextLine) VALUES( ?, ?, ?)",
                        (str(i), str(tmpValuesList[0]), str(tmpValuesList[1])))
            cnxn.commit()
    crsr.close()
    cnxn.close()


# testing calls
defsdict = (make_ana_def('config_file.xlsx', 'DI.xlsx'))
for line in defsdict:
    print(line, defsdict.get(line))
# mdbedit(defsdict)
